import os

from openpyxl import load_workbook

from docx import Document
from docxtpl import DocxTemplate

# Open files
from pathlib import Path

main_path = Path('./')
template_path = main_path / 'Contract_test-test.docx'
workbook_path = main_path / 'Contract_test-test.xlsx'

workbook = load_workbook(workbook_path)
template = DocxTemplate(template_path)
worksheet = workbook["Sheet1"]

to_fill_in = {'Company_Name' : None,
              'Your_Company_Slogan' : None,
              'Street_Address' : None,
              'City': None,
              'Date' : None,
              'Recipient_Name': None,
              'Recipient_Company_Name' : None,
              'Recipient_Street_Address' : None,
              'Recipient-City' : None,
              'Identification_Number' : None,
              }

# Set the minimum number of columns. This will be 2.
column = 2

# print out the maximum columns that are filled in in the excel file. This is to see how many iterations the code will need.
print(worksheet.max_column)

# Perform the following code block if the colomn amoumnt is less than the maximum column amount.
while column <= worksheet.max_column:

    # Define the column index. This is a letter so you need to convert the column number to a letter (2+64) = B
    col_index = chr(column + 64)
    row_index = 1
    # Retrieve the values from excel document and store in dictionary you defined earlier on
    # For each key in the dictionary, we look up the value in the excel file and store it instead of "none" in the dictionary
    for key in to_fill_in:
        cell = '%s%i' % (col_index, row_index)
        to_fill_in[key] = worksheet[cell].value
        row_index += 1

    # Fill in all the keys defined in the word document using the dictionary.
    # The keys in de word document are identified by the {{}}symbols.
    template.render(to_fill_in)

    # Output the file to a docx document.
    filename = str(to_fill_in['Company_Name']) + '_draft.docx'
    filled_path = os.path.join(main_path, filename)
    template.save(filled_path)
    print("Done with %s" % str(to_fill_in['Company_Name']))
    column += 1